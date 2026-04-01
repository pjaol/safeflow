#!/usr/bin/env python3
"""
Clio Daye content pipeline — Excel → JSON converter.

Reads Content/clio_content.xlsx (one sheet per content type) and writes
Resources/Content/*.json for bundling into the app.

Usage:
    python3 scripts/csv_to_json.py          # convert all sheets
    python3 scripts/csv_to_json.py tips     # convert only the 'tips' sheet

Requirements:
    pip3 install openpyxl

Sheets in the workbook:
    tips        Phase tips shown on PhaseTipCard
    nudges      Pattern nudges and comfort suggestions
    signals     Severity signals (escalating patterns)
    resources   Bundled support directory (orgs, helplines)
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip3 install openpyxl")
    sys.exit(1)

REPO_ROOT   = Path(__file__).parent.parent
WORKBOOK    = REPO_ROOT / "Content" / "clio_content.xlsx"
OUTPUT_DIR  = REPO_ROOT / "Resources" / "Content"

ALL_SHEETS  = ["tips", "nudges", "signals", "resources", "insights"]

# ── Type coercion ──────────────────────────────────────────────────────────────

INT_FIELDS  = {"cycle_count_min", "avg_cycle_max", "avg_cycle_min",
               "variability_min", "period_length_min", "threshold_value"}
BOOL_FIELDS = {"dismissible"}
LIST_FIELDS = {"symptoms_any", "symptoms_all", "languages", "tags"}


def coerce(key: str, value):
    """Convert a raw cell value to the appropriate Python type."""
    # openpyxl returns native Python types for numbers/bools; normalise to str first
    if value is None:
        raw = ""
    elif isinstance(value, bool):
        raw = str(value).lower()
    elif isinstance(value, (int, float)):
        raw = str(int(value)) if isinstance(value, float) and value == int(value) else str(value)
    else:
        raw = str(value).strip()

    if key in BOOL_FIELDS:
        return raw.lower() in ("true", "1", "yes")

    if key in INT_FIELDS:
        if raw == "":
            return None
        try:
            return int(float(raw))
        except ValueError:
            raise ValueError(f"Expected integer for field '{key}', got: {repr(raw)}")

    if key in LIST_FIELDS:
        if raw == "":
            return []
        return [item.strip() for item in raw.split(",") if item.strip()]

    return raw if raw != "" else None


# ── Conversion ─────────────────────────────────────────────────────────────────

def convert_sheet(wb, sheet_name: str) -> int:
    if sheet_name not in wb.sheetnames:
        print(f"  SKIP  sheet '{sheet_name}' not found in workbook")
        return 0

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"  SKIP  sheet '{sheet_name}' is empty")
        return 0

    headers = [str(h).strip() if h is not None else None for h in rows[0]]

    records = []
    for row_num, row in enumerate(rows[1:], start=2):
        # Skip entirely blank rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        record = {}
        for key, value in zip(headers, row):
            if key is None:
                continue
            record[key] = coerce(key, value)

        if not record.get("id"):
            raise ValueError(f"Sheet '{sheet_name}' row {row_num}: missing required 'id' field")

        records.append(record)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    dst = OUTPUT_DIR / f"{sheet_name}.json"
    with open(dst, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"  OK    {sheet_name:12s} → {dst.relative_to(REPO_ROOT)}  ({len(records)} records)")
    return len(records)


# ── Validation ─────────────────────────────────────────────────────────────────

def validate(sheet_name: str):
    dst = OUTPUT_DIR / f"{sheet_name}.json"
    if not dst.exists():
        return

    with open(dst, encoding="utf-8") as f:
        records = json.load(f)

    errors = []
    for r in records:
        rid = r.get("id", "<missing>")
        if sheet_name in ("tips", "nudges", "signals"):
            if not r.get("title"):
                errors.append(f"{rid}: missing 'title'")
            if not r.get("body"):
                errors.append(f"{rid}: missing 'body'")
        elif sheet_name == "resources":
            if not r.get("name"):
                errors.append(f"{rid}: missing 'name'")
            if not r.get("url") and not r.get("phone"):
                errors.append(f"{rid}: needs at least a 'url' or 'phone'")
        elif sheet_name == "insights":
            if not r.get("symptom"):
                errors.append(f"{rid}: missing 'symptom'")
            if not r.get("phase"):
                errors.append(f"{rid}: missing 'phase'")
            if not r.get("prevalence"):
                errors.append(f"{rid}: missing 'prevalence'")

    if errors:
        print(f"\n  VALIDATION ERRORS in {sheet_name}.json:")
        for e in errors:
            print(f"    - {e}")
        sys.exit(1)


# ── Workbook bootstrap ─────────────────────────────────────────────────────────

SHEET_SCHEMAS = {
    "tips": [
        "id", "phase", "title", "body", "sf_symbol", "priority",
        "cycle_count_min", "symptoms_any", "symptoms_all", "mood", "source",
    ],
    "nudges": [
        "id", "type", "title", "body", "sf_symbol", "priority",
        "cycle_count_min", "avg_cycle_max", "avg_cycle_min",
        "variability_min", "period_length_min",
        "symptoms_any", "symptoms_all", "dismissible", "source",
    ],
    "signals": [
        "id", "title", "body", "sf_symbol", "priority",
        "cycle_count_min", "check_type", "threshold_value", "source",
    ],
    "resources": [
        "id", "category", "name", "short_description", "url", "phone",
        "region", "languages", "tags",
    ],
    "insights": [
        "id", "symptom", "phase", "prevalence", "percentage_string", "note", "valence",
    ],
}

TIPS_DATA = [
    ("tip.menstrual.heat",         "menstrual",  "Heat can help",                  "A warm compress or hot water bottle on your lower abdomen may ease cramp discomfort.",                                                                                               "thermometer.medium",              "medium", 0, "",                       "", "", ""),
    ("tip.menstrual.rest",         "menstrual",  "Rest is productive",              "Your body is doing real work this week. Lighter activity and extra sleep are a good call.",                                                                                         "moon.fill",                       "medium", 0, "",                       "", "", ""),
    ("tip.menstrual.iron",         "menstrual",  "Iron-rich foods",                 "Leafy greens, lentils, and beans can help replenish iron lost during your period.",                                                                                                  "leaf.fill",                       "low",    0, "",                       "", "", ""),
    ("tip.menstrual.ibuprofen",    "menstrual",  "Ibuprofen timing",                "Anti-inflammatory pain relief works best taken with food at the first sign of cramps — not after pain peaks.",                                                                      "pill.fill",                       "medium", 0, "cramps",                 "", "", ""),
    ("tip.menstrual.hydration",    "menstrual",  "Stay hydrated",                   "Blood loss increases your need for fluids. Warm drinks can also ease cramping.",                                                                                                    "drop.fill",                       "low",    0, "",                       "", "", ""),
    ("tip.menstrual.movement",     "menstrual",  "Gentle movement can help",        "Light walking or gentle yoga may reduce cramp intensity for some people — listen to your body.",                                                                                    "figure.walk",                     "low",    0, "cramps",                 "", "", ""),
    ("tip.menstrual.bloating",     "menstrual",  "Bloating is common",              "Hormonal shifts cause water retention at the start of your period. It usually eases within a few days.",                                                                            "circle.dashed",                   "low",    0, "bloating",               "", "", ""),
    ("tip.menstrual.fatigue",      "menstrual",  "Fatigue is real",                 "The drop in oestrogen and progesterone at menstruation is physically demanding. Fatigue now is not laziness.",                                                                      "battery.25percent",               "medium", 0, "fatigue",                "", "", ""),
    ("tip.follicular.energy",      "follicular", "Energy is building",              "Oestrogen rises during this phase — many people find focus and motivation easier right now.",                                                                                       "arrow.up.circle.fill",            "medium", 0, "",                       "", "", ""),
    ("tip.follicular.new_things",  "follicular", "Good time for new things",        "Starting a new project or learning something new tends to feel easier in the follicular phase.",                                                                                    "brain.head.profile",              "low",    0, "",                       "", "", ""),
    ("tip.follicular.movement",    "follicular", "Movement feels easier",           "Higher oestrogen supports muscle recovery — a good window for more intense activity if that works for you.",                                                                        "figure.run",                      "low",    0, "",                       "", "", ""),
    ("tip.follicular.gut",         "follicular", "Support your gut",                "Fermented foods like yogurt and kimchi support the gut microbiome which influences hormone processing.",                                                                            "fork.knife",                      "low",    0, "",                       "", "", ""),
    ("tip.follicular.skin",        "follicular", "Skin often clears",               "Oestrogen tends to reduce sebum production in the follicular phase — many people notice clearer skin.",                                                                             "allergens",                       "low",    0, "acne",                   "", "", ""),
    ("tip.follicular.social",      "follicular", "Social energy rises",             "Many people feel more confident and outgoing as oestrogen climbs. A good time to schedule things you have been putting off.",                                                       "sun.max.fill",                    "low",    0, "",                       "", "", ""),
    ("tip.ovulatory.peak_energy",  "ovulatory",  "Peak energy window",              "Many people feel most social and energised around ovulation — a good time for presentations or important conversations.",                                                           "sun.max.fill",                    "medium", 0, "",                       "", "", ""),
    ("tip.ovulatory.temperature",  "ovulatory",  "Basal temperature rises",         "A slight rise in basal body temperature after ovulation is normal and confirms ovulation has occurred.",                                                                            "thermometer.medium",              "low",    0, "",                       "", "", ""),
    ("tip.ovulatory.hydration",    "ovulatory",  "Stay hydrated",                   "Hormonal shifts around ovulation can increase body temperature slightly. Extra water helps.",                                                                                       "drop.fill",                       "low",    0, "",                       "", "", ""),
    ("tip.ovulatory.body",         "ovulatory",  "Notice your body",                "Some people feel mild one-sided pelvic discomfort around ovulation (mittelschmerz). It is common and usually brief.",                                                              "figure.yoga",                     "low",    0, "mittelschmerz",           "", "", ""),
    ("tip.ovulatory.discharge",    "ovulatory",  "Discharge changes are normal",    "Cervical mucus becomes clearer and more elastic around ovulation. This is a normal hormonal sign.",                                                                                "drop.fill",                       "low",    0, "dischargeChanges",        "", "", ""),
    ("tip.ovulatory.communication","ovulatory",  "Communication feels easier",      "Verbal fluency and confidence often peak around ovulation. Difficult conversations may feel more manageable now.",                                                                  "bubble.left.and.bubble.right.fill","low",   0, "",                       "", "", ""),
    ("tip.luteal.mood",            "luteal",     "Mood may shift",                  "Progesterone rises then drops in the luteal phase. If your mood dips in the second half that is a recognised pattern.",                                                             "waveform.path.ecg",               "medium", 0, "",                       "", "", ""),
    ("tip.luteal.magnesium",       "luteal",     "Magnesium may help",              "Some research links magnesium intake to reduced PMS symptoms. Dark chocolate and nuts are decent sources.",                                                                         "heart.fill",                      "low",    0, "",                       "", ""),
    ("tip.luteal.sleep",           "luteal",     "Sleep changes are normal",        "Progesterone can affect sleep quality in the luteal phase. A cooler room and consistent sleep times help.",                                                                         "bed.double.fill",                 "medium", 0, "insomnia",                "", "", ""),
    ("tip.luteal.bloating",        "luteal",     "Reduce salt if bloated",          "Bloating is common in the late luteal phase. Reducing salty foods and increasing water can ease it.",                                                                               "drop.halffull",                   "low",    0, "bloating",                "", "", ""),
    ("tip.luteal.brain_fog",       "luteal",     "Brain fog is hormonal",           "Difficulty concentrating in the luteal phase is linked to progesterone. It is not a reflection of your ability.",                                                                   "cloud.fill",                      "medium", 0, "brainFog",                "", "", ""),
    ("tip.luteal.cravings",        "luteal",     "Cravings have a cause",           "Food cravings in the luteal phase are linked to serotonin fluctuations. Eating regularly helps stabilise them.",                                                                    "fork.knife",                      "low",    0, "foodCravings",            "", "", ""),
    ("tip.luteal.pmdd",            "luteal",     "Severe symptoms deserve attention","If mood symptoms in the second half of your cycle significantly disrupt your life that pattern is called PMDD and is treatable.",                                                  "waveform.path.ecg",               "high",   3, "anxious,irritable,sad",   "", "", ""),
    ("tip.luteal.breast",          "luteal",     "Breast tenderness is common",     "Progesterone causes breast tissue to swell in the luteal phase. A supportive bra and reducing caffeine can help.",                                                                  "heart.fill",                      "low",    0, "breastTenderness",        "", "", ""),
]

NUDGES_DATA = [
    ("nudge.pattern.shortCycle",    "health",   "Your cycles are running short",       "Your average cycle is running short. Cycles under 21 days are worth mentioning to a doctor — it is a common and treatable pattern.",                                                    "calendar",                  "high",   3, 21,  "",  "",  "",  "",  "",  True, ""),
    ("nudge.pattern.longCycle",     "health",   "Your cycles are running long",        "Your average cycle is running long. Cycles over 35 days can have a few different causes — a doctor can help figure out what is going on.",                                               "calendar",                  "high",   3, "",  35,  "",  "",  "",  "",  True, ""),
    ("nudge.pattern.highVariability","health",  "Your cycle length varies a lot",      "Your cycles vary by more than a week from month to month. This is fairly common but worth a chat with a doctor if it has been going on a while.",                                        "chart.line.uptrend.xyaxis", "medium", 3, "",  "",  7,   "",  "",  "",  True, ""),
    ("nudge.pattern.longPeriod",    "health",   "Your periods are running long",       "Periods consistently over 7 days are worth mentioning to a doctor. There are several common treatable causes.",                                                                          "stethoscope",               "medium", 3, "",  "",  "",  7,   "",  "",  True, ""),
    ("nudge.comfort.crampsHeavy",   "comfort",  "You have been logging cramps",        "A few things that can help: heat on your lower abdomen, ibuprofen taken with food at the first sign (not after), and magnesium-rich foods like dark chocolate and nuts.",                "thermometer.medium",        "medium", 2, "",  "",  "",  "",  "cramps", "", True, ""),
    ("nudge.comfort.fatigue",       "comfort",  "Persistent fatigue this phase",       "Fatigue around your period is common. Iron-rich foods, staying hydrated, and not skipping sleep all help. If it feels disabling it is worth mentioning to your doctor.",                 "battery.25percent",         "medium", 2, "",  "",  "",  "",  "fatigue", "", True, ""),
    ("nudge.comfort.bloating",      "comfort",  "Bloating logged consistently",        "Hormonal bloating before and during your period is very common. Reducing salt, eating smaller meals, and staying hydrated can help ease it.",                                             "circle.dashed",             "low",    2, "",  "",  "",  "",  "bloating", "", True, ""),
    ("nudge.comfort.insomnia",      "comfort",  "Sleep disruption logged",             "Progesterone fluctuations can disrupt sleep in the luteal and menstrual phases. A consistent bedtime, cooler room, and avoiding screens before sleep all help.",                         "moon.stars.fill",           "medium", 2, "",  "",  "",  "",  "insomnia", "", True, ""),
    ("nudge.comfort.brainFog",      "comfort",  "Brain fog logged consistently",       "Difficulty concentrating around your period is linked to hormone shifts, not your capability. Breaking tasks into smaller steps and building in rest helps.",                             "cloud.fill",                "medium", 2, "",  "",  "",  "",  "brainFog", "", True, ""),
]

SIGNALS_DATA = [
    ("signal.crampsEscalating",    "Your cramp days are increasing",        "Period pain that gets worse cycle over cycle is worth mentioning to a doctor. It can have treatable causes including endometriosis.", "chart.line.uptrend.xyaxis", "high",   3, "cramps_escalating",    "", ""),
    ("signal.heavyFlow",           "You have been logging heavy flow",      "Consistently heavy periods are common but worth discussing with a doctor — especially if you are soaking through protection quickly or feeling unusually fatigued.",  "drop.triangle.fill",       "high",   3, "heavy_flow_majority",  "", ""),
    ("signal.crampsFollicular",    "Cramps outside your period",            "You often log cramps during your follicular phase not just your period. This is less typical and worth mentioning to a doctor — there are several possible causes.",   "waveform.path.ecg",        "medium", 3, "cramps_follicular",    "", ""),
    ("signal.severeSymptomCluster","Severe symptoms most cycles",           "You are consistently logging multiple high-intensity symptoms. This pattern is worth discussing with a doctor — there are well-understood and treatable causes.",       "stethoscope",              "high",   3, "severe_cluster",       "", ""),
]

RESOURCES_DATA = [
    ("resource.acog.menstrual",  "ob_gyn",        "ACOG — Menstrual Disorders",            "Patient FAQs on abnormal uterine bleeding, painful periods, and cycle irregularities from the American College of OB/GYNs.",                    "https://www.acog.org/womens-health/faqs/abnormal-uterine-bleeding", "",              "US",     "en",    "heavy_flow,cramps,irregular"),
    ("resource.acog.endo",       "ob_gyn",        "ACOG — Endometriosis",                  "Plain-language patient guide to endometriosis diagnosis and treatment from the American College of OB/GYNs.",                                    "https://www.acog.org/womens-health/faqs/endometriosis",             "",              "US",     "en",    "cramps,pelvic_pain,endometriosis"),
    ("resource.acog.pcos",       "ob_gyn",        "ACOG — PCOS",                           "Patient FAQ on polycystic ovary syndrome — symptoms, diagnosis, and management.",                                                               "https://www.acog.org/womens-health/faqs/polycystic-ovary-syndrome", "",              "US",     "en",    "irregular,pcos"),
    ("resource.iapmd",           "pmdd",          "IAPMD — Premenstrual Disorders",        "The primary resource for PMDD and PME information, peer support, and provider directory.",                                                      "https://iapmd.org",                                                 "",              "global", "en",    "pmdd,mood,luteal"),
    ("resource.iapmd.crisis",    "pmdd",          "IAPMD — Crisis Support",               "IAPMD crisis support resources for people experiencing severe PMDD symptoms.",                                                                   "https://iapmd.org/crisis-support",                                  "",              "global", "en",    "pmdd,crisis,mood"),
    ("resource.efa",             "endometriosis", "Endometriosis Foundation of America",   "Patient resources, symptom diary tools, and information on finding a specialist for endometriosis.",                                             "https://www.endofound.org",                                         "",              "US",     "en",    "cramps,endometriosis,pelvic_pain"),
    ("resource.endouk",          "endometriosis", "Endometriosis UK",                      "UK patient charity for endometriosis — helpline, local support groups, and information on specialist centre networks.",                          "https://www.endometriosis-uk.org",                                  "0808 808 2227", "UK",     "en",    "cramps,endometriosis,pelvic_pain"),
    ("resource.verity",          "pcos",          "Verity — PCOS UK",                     "UK charity for people with PCOS — information, peer support, and guidance on getting a diagnosis.",                                             "https://www.verity-pcos.org.uk",                                    "",              "UK",     "en",    "pcos,irregular"),
    ("resource.pcosa",           "pcos",          "PCOS Awareness Association",            "US non-profit providing information and community for people with PCOS.",                                                                        "https://www.pcosaa.org",                                            "",              "US",     "en",    "pcos,irregular"),
    ("resource.nhs.periods",     "gp",            "NHS — Period Problems",                 "NHS guidance on heavy periods, painful periods, irregular periods, and when to see a GP.",                                                       "https://www.nhs.uk/conditions/periods/",                            "",              "UK",     "en",    "heavy_flow,cramps,irregular"),
    ("resource.nhs.endo",        "gp",            "NHS — Endometriosis",                   "NHS patient information on endometriosis symptoms, diagnosis, and treatment pathways.",                                                          "https://www.nhs.uk/conditions/endometriosis/",                      "",              "UK",     "en",    "cramps,endometriosis"),
    ("resource.nhs.pcos",        "gp",            "NHS — PCOS",                            "NHS patient information on polycystic ovary syndrome.",                                                                                         "https://www.nhs.uk/conditions/polycystic-ovary-syndrome-pcos/",     "",              "UK",     "en",    "pcos,irregular"),
    ("resource.medlineplus.mens", "gp",           "MedlinePlus — Menstruation",            "NIH plain-language health information on menstrual cycles and disorders. Public domain.",                                                        "https://medlineplus.gov/menstruation.html",                         "",              "US",     "en",    "general,heavy_flow,cramps,irregular"),
    ("resource.medlineplus.endo", "gp",           "MedlinePlus — Endometriosis",           "NIH plain-language information on endometriosis.",                                                                                              "https://medlineplus.gov/endometriosis.html",                        "",              "US",     "en",    "cramps,endometriosis"),
    ("resource.medlineplus.pmdd", "gp",           "MedlinePlus — PMDD",                   "NIH plain-language information on premenstrual dysphoric disorder.",                                                                             "https://medlineplus.gov/premenstrualdysphoricdisorder.html",        "",              "US",     "en",    "pmdd,mood,luteal"),
    ("resource.rainn",           "crisis",        "RAINN — Sexual Assault Hotline",        "National Sexual Assault Hotline — confidential support 24/7.",                                                                                  "https://hotline.rainn.org",                                         "1-800-656-4673","US",     "en",    "crisis,safety"),
    ("resource.ndvh",            "crisis",        "National Domestic Violence Hotline",    "24/7 confidential support for people experiencing domestic violence.",                                                                           "https://www.thehotline.org",                                        "1-800-799-7233","US",     "en",    "crisis,safety,dv"),
    ("resource.refuge",          "crisis",        "Refuge — Domestic Abuse",               "UK national charity providing support for women and children experiencing domestic violence.",                                                   "https://refuge.org.uk",                                             "0808 2000 247", "UK",     "en",    "crisis,safety,dv"),
    ("resource.crisis_text",     "mental_health", "Crisis Text Line",                      "Free 24/7 text-based mental health crisis support. Text HOME to 741741.",                                                                       "https://www.crisistextline.org",                                    "741741",        "US",     "en",    "crisis,mental_health,mood"),
    ("resource.samaritans",      "mental_health", "Samaritans",                            "UK free 24/7 emotional support helpline.",                                                                                                       "https://www.samaritans.org",                                        "116 123",       "UK",     "en",    "crisis,mental_health,mood"),
    ("resource.mind",            "mental_health", "Mind — Mental Health Charity",          "UK charity providing mental health information and local support.",                                                                              "https://www.mind.org.uk",                                           "",              "UK",     "en",    "mental_health,mood,pmdd"),
    ("resource.nami",            "mental_health", "NAMI — Mental Health Support",          "National Alliance on Mental Illness — information and support for mental health conditions in the US.",                                          "https://www.nami.org",                                              "1-800-950-6264","US",     "en",    "mental_health,mood"),
]

INSIGHTS_DATA = [
    # id                                 symptom             phase         prevalence      percentage_string   note                                                                                                                                 valence
    ("insight.menstrual.cramps",         "cramps",           "menstrual",  "very_common",  "around 70%",       "Cramps are one of the most reported period symptoms. Pain that stops you doing normal activities is worth discussing with a doctor.", ""),
    ("insight.menstrual.bloating",       "bloating",         "menstrual",  "very_common",  "around 65%",       "",                                                                                                                                  ""),
    ("insight.menstrual.backPain",       "backPain",         "menstrual",  "common",       "around 45%",       "",                                                                                                                                  ""),
    ("insight.menstrual.fatigue",        "fatigue",          "menstrual",  "very_common",  "around 60%",       "Iron loss during menstruation contributes to fatigue for many people.",                                                              ""),
    ("insight.menstrual.headache",       "headache",         "menstrual",  "common",       "around 35%",       "Hormone shifts at the start of the cycle can trigger headaches.",                                                                   ""),
    ("insight.menstrual.nausea",         "nausea",           "menstrual",  "fairly_common","around 20%",       "",                                                                                                                                  ""),
    ("insight.menstrual.breastTenderness","breastTenderness","menstrual",  "fairly_common","around 25%",       "",                                                                                                                                  ""),
    ("insight.menstrual.insomnia",       "insomnia",         "menstrual",  "fairly_common","around 20%",       "",                                                                                                                                  ""),
    ("insight.menstrual.foodCravings",   "foodCravings",     "menstrual",  "common",       "around 50%",       "",                                                                                                                                  ""),
    ("insight.menstrual.brainFog",       "brainFog",         "menstrual",  "fairly_common","around 25%",       "",                                                                                                                                  ""),
    ("insight.menstrual.acne",           "acne",             "menstrual",  "fairly_common","around 30%",       "",                                                                                                                                  ""),
    ("insight.follicular.fatigue",       "fatigue",          "follicular", "less_common",  "around 15%",       "Fatigue is less typical in the follicular phase as oestrogen rises. If it is persistent, worth noting.",                            ""),
    ("insight.follicular.headache",      "headache",         "follicular", "less_common",  "around 10%",       "",                                                                                                                                  ""),
    ("insight.follicular.acne",          "acne",             "follicular", "fairly_common","around 20%",       "",                                                                                                                                  ""),
    ("insight.follicular.bloating",      "bloating",         "follicular", "less_common",  "around 15%",       "",                                                                                                                                  ""),
    ("insight.follicular.cramps",        "cramps",           "follicular", "rare",         "around 5-10%",     "Cramps outside your period are less typical. If this is consistent it is worth mentioning to a doctor.",                            ""),
    ("insight.follicular.highEnergy",    "highEnergy",       "follicular", "common",       "around 55%",       "Many people notice a natural energy lift during the follicular phase.",                                                              ""),
    ("insight.follicular.brainFog",      "brainFog",         "follicular", "less_common",  "around 10%",       "",                                                                                                                                  ""),
    ("insight.ovulatory.mittelschmerz",  "mittelschmerz",    "ovulatory",  "common",       "around 40%",       "A brief one-sided ache or twinge around ovulation is very common and harmless.",                                                   ""),
    ("insight.ovulatory.dischargeChanges","dischargeChanges","ovulatory",  "very_common",  "around 70%",       "Discharge changes around ovulation are a normal hormonal signal.",                                                                  ""),
    ("insight.ovulatory.bloating",       "bloating",         "ovulatory",  "fairly_common","around 25%",       "",                                                                                                                                  ""),
    ("insight.ovulatory.headache",       "headache",         "ovulatory",  "fairly_common","around 20%",       "The oestrogen peak before ovulation can trigger headaches in some people.",                                                         ""),
    ("insight.ovulatory.breastTenderness","breastTenderness","ovulatory",  "fairly_common","around 30%",       "",                                                                                                                                  ""),
    ("insight.ovulatory.highEnergy",     "highEnergy",       "ovulatory",  "very_common",  "around 60%",       "",                                                                                                                                  ""),
    ("insight.luteal.breastTenderness",  "breastTenderness", "luteal",     "very_common",  "around 60%",       "",                                                                                                                                  ""),
    ("insight.luteal.bloating",          "bloating",         "luteal",     "very_common",  "around 65%",       "",                                                                                                                                  ""),
    ("insight.luteal.foodCravings",      "foodCravings",     "luteal",     "very_common",  "around 60%",       "",                                                                                                                                  ""),
    ("insight.luteal.fatigue",           "fatigue",          "luteal",     "common",       "around 45%",       "Energy often dips in the late luteal phase as progesterone peaks.",                                                                 ""),
    ("insight.luteal.insomnia",          "insomnia",         "luteal",     "common",       "around 35%",       "Progesterone fluctuations can affect sleep quality.",                                                                               ""),
    ("insight.luteal.brainFog",          "brainFog",         "luteal",     "common",       "around 40%",       "",                                                                                                                                  ""),
    ("insight.luteal.acne",              "acne",             "luteal",     "common",       "around 45%",       "",                                                                                                                                  ""),
    ("insight.luteal.headache",          "headache",         "luteal",     "common",       "around 35%",       "Pre-menstrual headaches caused by the oestrogen drop before your period are very common.",                                          ""),
    ("insight.luteal.cramps",            "cramps",           "luteal",     "fairly_common","around 20%",       "Some cramping in the late luteal phase is common as the uterus prepares.",                                                          ""),
    ("insight.luteal.appetiteChanges",   "appetiteChanges",  "luteal",     "common",       "around 50%",       "",                                                                                                                                  ""),
    ("insight.luteal.backPain",          "backPain",         "luteal",     "fairly_common","around 25%",       "",                                                                                                                                  ""),
    ("insight.mood.luteal.negative",     "mood",             "luteal",     "very_common",  "around 75%",       "Mood changes in the luteal phase are one of the most reported cycle experiences.",                                                  "negative"),
    ("insight.mood.menstrual.negative",  "mood",             "menstrual",  "common",       "around 50%",       "",                                                                                                                                  "negative"),
    ("insight.mood.follicular.positive", "mood",             "follicular", "common",       "around 55%",       "Rising oestrogen in this phase often lifts mood and energy.",                                                                       "positive"),
    ("insight.mood.ovulatory.positive",  "mood",             "ovulatory",  "common",       "around 55%",       "Rising oestrogen in this phase often lifts mood and energy.",                                                                       "positive"),
]

SHEET_DATA = {
    "tips":      TIPS_DATA,
    "nudges":    NUDGES_DATA,
    "signals":   SIGNALS_DATA,
    "resources": RESOURCES_DATA,
    "insights":  INSIGHTS_DATA,
}


def create_workbook():
    """Create clio_content.xlsx with all content pre-populated."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_font  = Font(bold=True, color="FFFFFF")
    header_fills = {
        "tips":      PatternFill("solid", fgColor="4A90D9"),
        "nudges":    PatternFill("solid", fgColor="7B5EA7"),
        "signals":   PatternFill("solid", fgColor="D94F5C"),
        "resources": PatternFill("solid", fgColor="2E9E6B"),
        "insights":  PatternFill("solid", fgColor="E57C4A"),
    }

    for sheet_name in ALL_SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        headers = SHEET_SCHEMAS[sheet_name]
        data    = SHEET_DATA[sheet_name]

        # Header row
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fills[sheet_name]
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_num, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)

        # Auto-width columns (capped at 60)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        # Freeze header row
        ws.freeze_panes = "A2"

    WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK)
    print(f"  Created {WORKBOOK.relative_to(REPO_ROOT)}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    targets = sys.argv[1:] if len(sys.argv) > 1 else ALL_SHEETS
    unknown = [t for t in targets if t not in ALL_SHEETS]
    if unknown:
        print(f"Unknown sheet(s): {unknown}. Valid: {ALL_SHEETS}")
        sys.exit(1)

    print("Clio Daye content pipeline")
    print(f"  Workbook: {WORKBOOK.relative_to(REPO_ROOT)}")
    print(f"  Output:   {OUTPUT_DIR.relative_to(REPO_ROOT)}/")
    print()

    if not WORKBOOK.exists():
        print("  Workbook not found — creating from seed data...")
        create_workbook()
        print()

    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)

    total = 0
    for name in targets:
        total += convert_sheet(wb, name)
        validate(name)

    print(f"\n  {total} records across {len(targets)} sheet(s). Commit Resources/Content/ to update the app.")


if __name__ == "__main__":
    main()
